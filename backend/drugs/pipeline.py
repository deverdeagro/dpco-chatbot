"""
Reusable end-to-end classification pipeline over a workbook, shared by the
`classify_products` management command and the report-upload API.

    workbook -> map_columns (LLM) -> normalize (LLM) -> classify (rules) -> write

`classify_workbook` returns the per-row results and a summary, and writes a copy
of the workbook with three new columns (Normalized Composition, Classification,
Reason). The original input file is never modified.
"""

import json

import openpyxl

from drugs.normalize import normalize_composition
from drugs.classify import build_nlem_index, classify_composition
from drugs.resolver import make_resolver
from drugs.sheet_mapper import map_columns

NEW_COLUMNS = ["Normalized Composition", "Classification", "Reason"]


def classify_workbook(input_path, output_path, sheet=None, on_progress=None):
    """
    Classify every product row in a sheet.

    Args:
      input_path:  path to the .xlsx to read.
      output_path: path to write the classified copy to.
      sheet:       sheet name; defaults to the first sheet.
      on_progress: optional callable(processed, total, rows, counts) called as
                   rows finish, so callers can persist partial results live.

    Returns: (rows, summary)
      rows = [{"row", "brand", "composition", "classification", "reason"}, ...]
      summary = {"total": int, "counts": {label: n}, "sheet": str}
    """
    wb = openpyxl.load_workbook(input_path, data_only=True)
    sheet = sheet or wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]

    mapping = map_columns(ws)
    header_row = mapping["header_row"]
    comp_col = mapping["columns"]["composition"]
    brand_col = mapping["columns"]["brand"]
    dosage_col = mapping["columns"]["dosage_form"]

    # Write the new column headers after the existing columns.
    first_new = ws.max_column + 1
    for offset, title in enumerate(NEW_COLUMNS):
        ws.cell(row=header_row, column=first_new + offset, value=title)

    index = build_nlem_index()
    resolver = make_resolver(index)

    # Count data rows up front so progress has a denominator.
    data_rows = [
        (r, row)
        for r, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                                start=header_row + 1)
        if comp_col is not None and row[comp_col] is not None and str(row[comp_col]).strip()
    ]
    total = len(data_rows)

    rows, counts = [], {}
    for processed, (r, row) in enumerate(data_rows, start=1):
        composition = str(row[comp_col]).strip()
        # A single unparseable/odd row must never abort the whole batch:
        # record it as an "error" classification and keep going.
        try:
            norm = normalize_composition(composition)
            if not norm.get("dosage_form") and dosage_col is not None and row[dosage_col]:
                norm["dosage_form"] = str(row[dosage_col]).strip()
            result = classify_composition(norm, index, resolver=resolver)
            label = result["classification"]
            reason = result["reason"]
        except Exception as exc:  # noqa: BLE001
            norm = {"error": str(exc)}
            label = "error"
            reason = f"Could not classify this row: {exc}"

        counts[label] = counts.get(label, 0) + 1

        ws.cell(row=r, column=first_new + 0, value=json.dumps(norm, ensure_ascii=False))
        ws.cell(row=r, column=first_new + 1, value=label)
        ws.cell(row=r, column=first_new + 2, value=reason)

        brand = row[brand_col] if brand_col is not None else None
        rows.append({
            "row": r,
            "brand": str(brand).strip() if brand else "",
            "composition": composition,
            "classification": label,
            "reason": reason,
        })

        if on_progress:
            on_progress(processed, total, rows, counts)

    wb.save(output_path)
    return rows, {"total": total, "counts": counts, "sheet": sheet}
