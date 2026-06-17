import re
import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from drugs.models import CeilingPrice

SO_PREFIX_RE = re.compile(r"^\s*S\.?\s*O\.?\s*", re.IGNORECASE)

# Sheets to skip entirely (not ceiling price data)
SKIP_SHEETS = {"Retail Price", "WPI"}


def load_wpi_rates(wb):
    """Read year→WPI rate from the WPI sheet. Returns {year_str: Decimal}."""
    if "WPI" not in wb.sheetnames:
        return {}
    rates = {}
    for row in wb["WPI"].iter_rows(min_row=1, values_only=True):
        year_val, wpi_val = row[0], row[1]
        if year_val is None or wpi_val is None:
            continue
        try:
            year = str(int(float(year_val)))
            rates[year] = Decimal(str(wpi_val))
        except (ValueError, InvalidOperation):
            continue
    return rates


def financial_year(year_str):
    y = int(year_str)
    return f"{y}-{(y + 1) % 100:02d}"


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def clean_so_number(value):
    text = clean_text(value)
    return SO_PREFIX_RE.sub("", text).strip()


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = clean_text(value)
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_price(value):
    if value is None:
        return None
    if isinstance(value, str):
        if value.strip().startswith("="):
            return None
        try:
            return Decimal(value.strip())
        except InvalidOperation:
            return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


class Command(BaseCommand):
    help = "Import DPCO ceiling prices from the Excel workbook into the database"

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the .xlsx file")
        parser.add_argument("--clear", action="store_true", help="Delete existing records before import")

    def handle(self, *args, **options):
        filepath = options["file"]

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {filepath}")

        wpi_rates = load_wpi_rates(wb)
        if not wpi_rates:
            raise CommandError("WPI sheet not found or empty — cannot determine WPI rates.")
        self.stdout.write(f"Loaded WPI rates for years: {', '.join(sorted(wpi_rates))}")

        if options["clear"]:
            deleted, _ = CeilingPrice.objects.all().delete()
            self.stdout.write(f"Cleared {deleted} existing records.")

        total_imported = 0
        total_skipped = 0

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            # Determine base year for WPI lookup
            # "2017 Post GST" maps to year "2017"
            base_year = sheet_name.split()[0]
            if base_year not in wpi_rates:
                self.stdout.write(self.style.WARNING(
                    f"Sheet '{sheet_name}': no WPI rate for year '{base_year}', skipping."
                ))
                continue

            wpi_rate = wpi_rates[base_year]
            fy = financial_year(base_year)
            combine_dosage_strength = (sheet_name == "2016")

            ws = wb[sheet_name]
            records = []
            skipped = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                so_number = clean_so_number(row[1])
                so_date = parse_date(row[2])
                medicine_name = clean_text(row[3])

                if combine_dosage_strength:
                    # 2016 layout: Strength(col4), Dosage(col5), Unit(col6), Ceiling Price(col7), Effective From(col8)
                    strength = clean_text(row[4])
                    dosage = clean_text(row[5])
                    dosage_form_and_strength = " ".join(p for p in [dosage, strength] if p)
                    unit = clean_text(row[6])
                    ceiling_price = parse_price(row[7])
                    effective_from = parse_date(row[8])
                else:
                    # 2017+ layout: Dosage form and Strength(col4), Unit(col5), Ceiling price(col6), Effective From(col7)
                    dosage_form_and_strength = clean_text(row[4])
                    unit = clean_text(row[5])
                    ceiling_price = parse_price(row[6])
                    effective_from = parse_date(row[7])

                if not medicine_name or ceiling_price is None or effective_from is None:
                    skipped += 1
                    continue

                records.append(CeilingPrice(
                    so_number=so_number,
                    so_date=so_date,
                    medicine_name=medicine_name,
                    dosage_form_and_strength=dosage_form_and_strength,
                    unit=unit,
                    ceiling_price=ceiling_price,
                    wpi_rate=wpi_rate,
                    effective_from=effective_from,
                    financial_year=fy,
                    source_sheet=sheet_name,
                    row_number=row_idx,
                ))

            CeilingPrice.objects.bulk_create(records)
            self.stdout.write(self.style.SUCCESS(
                f"[{sheet_name}] Imported {len(records)} records. Skipped {skipped} blank/invalid rows."
            ))
            total_imported += len(records)
            total_skipped += skipped

        self.stdout.write(self.style.SUCCESS(
            f"\nTotal imported: {total_imported}. Total skipped: {total_skipped}."
        ))
