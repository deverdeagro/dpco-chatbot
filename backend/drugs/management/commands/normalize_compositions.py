"""
Run the composition normalizer over a client price sheet and print the
structured output, so the result can be eyeballed against the source.

    python manage.py normalize_compositions /path/to/Form-5-SampleFilings.xlsx
    python manage.py normalize_compositions <file> --sheet "New Products" --limit 5

This is a demonstration / inspection command for step 1 of the pipeline.
It only reads the sheet; it writes nothing to the database.
"""

import json

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from drugs.normalize import normalize_composition

# Header labels we know how to find the composition under, in priority order.
COMPOSITION_HEADERS = [
    "Drug (Composition DCA)",
    "Drug (Composition with Strength)",
    "Composition with strength",
    "Composition",
]
BRAND_HEADERS = ["Brand Name", "Product Name"]


def _find_header_row(ws):
    """Return (row_index, {normalized_header: col_index}) for the header row."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        labels = {str(v).strip(): i for i, v in enumerate(row) if v is not None}
        if any(h in labels for h in COMPOSITION_HEADERS):
            return r, labels
    return None, {}


def _pick(labels, candidates):
    for c in candidates:
        if c in labels:
            return labels[c]
    return None


class Command(BaseCommand):
    help = "Normalize composition strings from a price sheet into structured JSON."

    def add_arguments(self, parser):
        parser.add_argument("xlsx", help="Path to the .xlsx price sheet")
        parser.add_argument("--sheet", default="New Products", help="Sheet name")
        parser.add_argument("--limit", type=int, default=0, help="Max rows (0 = all)")

    def handle(self, *args, **opts):
        try:
            wb = openpyxl.load_workbook(opts["xlsx"], read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise CommandError(f"Could not open workbook: {exc}")

        if opts["sheet"] not in wb.sheetnames:
            raise CommandError(
                f"Sheet {opts['sheet']!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[opts["sheet"]]

        header_row, labels = _find_header_row(ws)
        if header_row is None:
            raise CommandError("Could not locate a composition column in the sheet.")

        comp_col = _pick(labels, COMPOSITION_HEADERS)
        brand_col = _pick(labels, BRAND_HEADERS)

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            composition = row[comp_col] if comp_col is not None else None
            if composition is None or not str(composition).strip():
                continue

            brand = row[brand_col] if brand_col is not None else None
            result = normalize_composition(str(composition))

            self.stdout.write(self.style.MIGRATE_HEADING(str(brand or "—")))
            self.stdout.write(f"  source: {str(composition).strip()}")
            self.stdout.write(f"  parsed: {json.dumps(result, ensure_ascii=False)}")
            self.stdout.write("")

            count += 1
            if opts["limit"] and count >= opts["limit"]:
                break

        self.stdout.write(self.style.SUCCESS(f"Normalized {count} composition(s)."))
