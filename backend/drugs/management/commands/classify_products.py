"""
Run the full pipeline over a price sheet and write the result back as new
columns in a COPY of the workbook (the original is never modified).

    python manage.py classify_products ~/Desktop/DPCO/Form-5-SampleFilings.xlsx --sheet "MRP Rev"
    python manage.py classify_products <file> --sheet "MRP Rev" --limit 5 --out /tmp/out.xlsx

Pipeline per row:  composition -> normalize (LLM) -> classify (NLEM rules)
New columns added: "Normalized Composition", "Classification", "Reason".
"""

import json
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openai import APIConnectionError, APITimeoutError

from drugs.normalize import normalize_composition
from drugs.classify import build_nlem_index, classify_composition
from drugs.resolver import make_resolver
from drugs.sheet_mapper import map_columns

NEW_COLUMNS = ["Normalized Composition", "Classification", "Reason"]


class Command(BaseCommand):
    help = "Classify each product in a sheet as scheduled / new drug against NLEM."

    def add_arguments(self, parser):
        parser.add_argument("xlsx")
        parser.add_argument("--sheet", default="MRP Rev")
        parser.add_argument("--limit", type=int, default=0)
        parser.add_argument("--out", default=None, help="Output path (default: <name>-classified.xlsx)")

    def handle(self, *args, **opts):
        path = Path(opts["xlsx"]).expanduser()
        out = Path(opts["out"]).expanduser() if opts["out"] else path.with_name(path.stem + "-classified.xlsx")

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise CommandError(f"Could not open workbook: {exc}")
        if opts["sheet"] not in wb.sheetnames:
            raise CommandError(f"Sheet {opts['sheet']!r} not found. Available: {wb.sheetnames}")
        ws = wb[opts["sheet"]]

        # Let the LLM figure out which columns are which, for any layout.
        self.stdout.write("Mapping sheet columns...")
        try:
            mapping = map_columns(ws)
        except APIConnectionError:
            raise CommandError(
                "Cannot reach the LLM (Ollama). Start it with `ollama serve` "
                f"and ensure model '{settings.OLLAMA_MODEL}' is pulled, then re-run."
            )
        header_row = mapping["header_row"]
        comp_col = mapping["columns"]["composition"]
        brand_col = mapping["columns"]["brand"]
        dosage_col = mapping["columns"]["dosage_form"]
        self.stdout.write(f"  header row {header_row}, composition col {comp_col}, "
                          f"brand col {brand_col}, dosage col {dosage_col}")

        # Write the new column headers after the existing columns.
        first_new = ws.max_column + 1
        for offset, title in enumerate(NEW_COLUMNS):
            ws.cell(row=header_row, column=first_new + offset, value=title)

        self.stdout.write("Building NLEM index...")
        index = build_nlem_index()
        resolver = make_resolver(index)  # LLM fallback for names rules can't match

        counts, processed = {}, 0
        for r, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                                start=header_row + 1):
            composition = row[comp_col] if comp_col is not None else None
            if composition is None or not str(composition).strip():
                continue

            try:
                norm = normalize_composition(str(composition))
            except (APIConnectionError, APITimeoutError):
                raise CommandError(
                    "Cannot reach the LLM (Ollama) or it timed out. Start it with "
                    f"`ollama serve` and ensure model '{settings.OLLAMA_MODEL}' is "
                    "pulled, then re-run."
                )
            # Backfill dosage form from the sheet's own column if missing.
            if not norm.get("dosage_form") and dosage_col is not None and row[dosage_col]:
                norm["dosage_form"] = str(row[dosage_col]).strip()

            result = classify_composition(norm, index, resolver=resolver)
            counts[result["classification"]] = counts.get(result["classification"], 0) + 1

            ws.cell(row=r, column=first_new + 0, value=json.dumps(norm, ensure_ascii=False))
            ws.cell(row=r, column=first_new + 1, value=result["classification"])
            ws.cell(row=r, column=first_new + 2, value=result["reason"])

            brand = row[brand_col] if brand_col is not None else "—"
            self.stdout.write(f"  {str(brand or '—').ljust(24)} -> {result['classification']}")

            processed += 1
            if opts["limit"] and processed >= opts["limit"]:
                break

        wb.save(out)
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"Classified {processed} product(s): {dict(counts)}"))
        self.stdout.write(self.style.SUCCESS(f"Saved -> {out}"))
